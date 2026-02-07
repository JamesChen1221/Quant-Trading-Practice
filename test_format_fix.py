"""
測試格式修正
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import pandas as pd

input_file = "量化交易.xlsx"
sheet_name = "資料庫"

print("="*60)
print("測試格式修正")
print("="*60)

# 讀取 Excel
df = pd.read_excel(input_file, sheet_name=sheet_name)

# 載入工作簿
wb = load_workbook(input_file)
ws = wb[sheet_name]

# 檢查第 2 行的格式
print("\n檢查第 2 行的格式:")
print("-"*60)

# 公司代碼欄位（參考格式）
ticker_col_idx = df.columns.get_loc('公司代碼') + 1
ref_cell = ws.cell(row=2, column=ticker_col_idx)

print(f"參考儲存格（公司代碼）:")
print(f"  字型: {ref_cell.font.name}")
print(f"  大小: {ref_cell.font.size}")
print(f"  粗體: {ref_cell.font.bold}")
print(f"  對齊: {ref_cell.alignment.vertical}")

# 檢查幾個自動填入的欄位
test_cols = ['5天 RSI 序列', '*昨日收盤價', '*開盤價格']
print(f"\n檢查自動填入的欄位:")
print("-"*60)

for col_name in test_cols:
    if col_name in df.columns:
        col_idx = df.columns.get_loc(col_name) + 1
        cell = ws.cell(row=2, column=col_idx)
        
        print(f"\n{col_name}:")
        print(f"  值: {cell.value}")
        print(f"  字型: {cell.font.name}")
        print(f"  大小: {cell.font.size}")
        print(f"  對齊: {cell.alignment.vertical}")
        
        # 檢查是否與參考格式一致
        if cell.font.name == ref_cell.font.name and cell.font.size == ref_cell.font.size:
            print(f"  ✓ 格式正確")
        else:
            print(f"  ✗ 格式不一致")

wb.close()

print("\n" + "="*60)
print("建議:")
print("  如果格式還是不一致，可能需要:")
print("  1. 檢查 Excel 是否有條件格式")
print("  2. 檢查是否有儲存格樣式")
print("  3. 手動設定一次格式後再執行程式")
print("="*60)
