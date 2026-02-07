import yfinance as yf
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
import numpy as np

def calculate_intraday_prices(ticker, trade_date):
    """
    計算盤中價格指標（僅限最近 7 天）
    
    指標：
    - 開盤價
    - 10分鐘最低價（開盤後 10 分鐘內的最低價）
    - 1.5小時最高價（開盤 10 分鐘後 ~ 1.5 小時的最高價）
    - 最高價前的最低價（開盤 10 分鐘後 ~ 1.5小時最高價時間點的最低價）
    
    參數:
        ticker: 股票代號
        trade_date: 交易日期（美國時間）
    
    返回:
        dict: 包含各項價格指標，如果無法獲取則返回 None
    """
    try:
        # 確保日期格式正確
        if isinstance(trade_date, str):
            trade_date = pd.to_datetime(trade_date)
        
        # 檢查是否在最近 7 天內
        days_ago = (datetime.now() - trade_date).days
        if days_ago > 7:
            print(f"  ⚠ {trade_date.strftime('%Y-%m-%d')} 超過 7 天，無法獲取盤中數據（yfinance 限制）")
            return None
        
        # 下載分鐘級數據
        print(f"  正在下載 {ticker} 的盤中數據...")
        
        # 設定日期範圍（當天）
        start = trade_date.strftime("%Y-%m-%d")
        end = (trade_date + timedelta(days=1)).strftime("%Y-%m-%d")
        
        # 下載 1 分鐘數據
        df = yf.download(ticker, start=start, end=end, interval="1m", progress=False)
        
        if df.empty:
            print(f"  ⚠ 無法獲取 {ticker} 在 {start} 的盤中數據")
            return None
        
        # 處理多層索引
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        
        # 確保有必要的欄位
        if 'Open' not in df.columns or 'High' not in df.columns or 'Low' not in df.columns:
            print(f"  ⚠ 數據不完整")
            return None
        
        # 確保數據是一維的
        for col in ['Open', 'High', 'Low', 'Close']:
            if col in df.columns and df[col].ndim > 1:
                df[col] = df[col].iloc[:, 0]
        
        # 美股開盤時間：9:30 AM ET
        # 假設數據已經是 ET 時區
        
        # 1. 開盤價（第一根 K 棒的開盤價）
        open_price = df['Open'].iloc[0] if len(df) > 0 else None
        
        # 2. 10分鐘最低價（開盤後 10 分鐘內，即前 10 根 1 分鐘 K 棒）
        first_10_min = df.head(10)
        low_10min = first_10_min['Low'].min() if len(first_10_min) > 0 else None
        
        # 3. 1.5小時最高價（開盤 10 分鐘後 ~ 1.5 小時，即第 11 根到第 90 根 K 棒）
        # 10 分鐘 = 10 根，1.5 小時 = 90 根
        after_10min_to_90min = df.iloc[10:90] if len(df) > 10 else pd.DataFrame()
        
        if len(after_10min_to_90min) > 0:
            high_90min = after_10min_to_90min['High'].max()
            # 找到最高價的時間點（索引位置）
            high_90min_idx = after_10min_to_90min['High'].idxmax()
            
            # 4. 最高價前的最低價（開盤 10 分鐘後 ~ 最高價時間點的最低價）
            # 找到最高價在原始 df 中的位置
            high_position = df.index.get_loc(high_90min_idx)
            
            # 如果最高價出現在第 11 根 K 棒（索引 10），表示開盤後立即達到高點
            if high_position == 10:
                print(f"  ⚠ 最高價出現在開盤後第 11 分鐘，設為最高價（開盤後立即達到高點）")
                low_before_high = high_90min
            elif high_position > 10:
                # 從第 11 根（索引 10）到最高價位置（包含）
                before_high = df.iloc[10:high_position+1]
                low_before_high = before_high['Low'].min() if len(before_high) > 0 else high_90min
            else:
                # 理論上不應該發生（因為我們從第 11 根開始找）
                print(f"  ⚠ 異常：最高價位置 {high_position} < 10")
                low_before_high = high_90min
        else:
            high_90min = None
            low_before_high = None
        
        # 檢查是否有足夠的數據
        if len(df) < 90:
            print(f"  ⚠ 數據不足 90 分鐘（只有 {len(df)} 分鐘）")
        
        return {
            "開盤價": round(open_price, 2) if open_price else None,
            "10分鐘最低價": round(low_10min, 2) if low_10min else None,
            "1.5小時最高價": round(high_90min, 2) if high_90min else None,
            "最高價前的最低價": round(low_before_high, 2) if low_before_high else None,
            "數據分鐘數": len(df)
        }
    
    except Exception as e:
        print(f"  ✗ 獲取盤中數據時發生錯誤: {str(e)}")
        return None

def calculate_price_distance(close_prices, current_price, days):
    """
    計算當前價格距離過去 N 天最高/最低價的百分比
    
    參數:
        close_prices: 收盤價序列
        current_price: 當前價格（通常是昨日收盤價）
        days: 回溯天數
    
    返回:
        dict: 包含距離最高價和最低價的百分比
    """
    if len(close_prices) < days:
        return None
    
    # 取最近 N 天的收盤價
    recent_prices = close_prices.tail(days)
    
    # 找出最高價和最低價
    highest = recent_prices.max()
    lowest = recent_prices.min()
    
    # 計算距離百分比
    distance_to_high = ((current_price - highest) / highest) * 100
    distance_to_low = ((current_price - lowest) / lowest) * 100
    
    return {
        "最高價": round(highest, 2),
        "最低價": round(lowest, 2),
        "距離最高價(%)": round(distance_to_high, 1),
        "距離最低價(%)": round(distance_to_low, 1)
    }

def calculate_rsi(close_prices, period=14):
    """計算 RSI 指標"""
    delta = close_prices.diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()
    rs = gain / loss
    rsi = 100 - (100 / (1 + rs))
    return rsi

def calculate_adx(high, low, close, period=14):
    """計算 ADX 指標"""
    # 計算 True Range
    high_low = high - low
    high_close = np.abs(high - close.shift())
    low_close = np.abs(low - close.shift())
    tr = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
    
    # 計算方向移動
    up_move = high - high.shift()
    down_move = low.shift() - low
    
    plus_dm = np.where((up_move > down_move) & (up_move > 0), up_move, 0)
    minus_dm = np.where((down_move > up_move) & (down_move > 0), down_move, 0)
    
    plus_dm = pd.Series(plus_dm, index=close.index)
    minus_dm = pd.Series(minus_dm, index=close.index)
    
    # 平滑化
    atr = tr.rolling(window=period).mean()
    plus_di = 100 * (plus_dm.rolling(window=period).mean() / atr)
    minus_di = 100 * (minus_dm.rolling(window=period).mean() / atr)
    
    # 計算 DX 和 ADX
    dx = 100 * np.abs(plus_di - minus_di) / (plus_di + minus_di)
    adx = dx.rolling(window=period).mean()
    
    return adx

def calculate_rsi_adx_sequences(ticker, start_date, days_5=5, days_30=30, days_180=120):
    """
    計算指定股票在開盤日期之前的 RSI、ADX 和價格距離（過去的資料）
    
    參數:
        ticker: 股票代碼 (例如: "NVDA")
        start_date: 開盤日期 (字串或 datetime)
        days_5: 5天序列長度
        days_30: 30天序列長度
        days_180: 6個月序列長度（約120個交易日）
    
    返回:
        dict: 包含 RSI、ADX 的序列和價格距離
    """
    try:
        # 確保日期格式正確
        if isinstance(start_date, str):
            start_date = pd.to_datetime(start_date)
        
        # 計算需要抓取的日期範圍（需要更多資料以計算 6 個月序列）
        fetch_start = start_date - timedelta(days=250)  # 約 8 個月前，確保有足夠交易日
        fetch_end = start_date + timedelta(days=1)
        
        # 下載股票資料
        print(f"  正在下載 {ticker} 的資料...")
        df = yf.download(ticker, start=fetch_start, end=fetch_end, progress=False)
        
        if df.empty:
            print(f"  警告: {ticker} 沒有資料")
            return None
        
        # 處理多層索引的情況
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        
        # 確保資料是一維的
        for col in ['Open', 'High', 'Low', 'Close', 'Volume']:
            if col in df.columns and df[col].ndim > 1:
                df[col] = df[col].iloc[:, 0]
        
        # 計算 RSI (14)
        df["RSI"] = calculate_rsi(df["Close"], period=14)
        
        # 計算 ADX (14)
        df["ADX"] = calculate_adx(df["High"], df["Low"], df["Close"], period=14)
        
        # 找到開盤日期當天或之前的資料
        df_before_start = df[df.index <= start_date]
        
        # 找到開盤日期的前一天（昨日）
        # 注意：這裡要找開盤日期之前的最後一個交易日
        df_before_open = df[df.index < start_date]
        
        if len(df_before_open) == 0:
            print(f"  警告: {ticker} 找不到開盤日期之前的資料")
            return None
        
        if len(df_before_start) < days_180:
            print(f"  警告: {ticker} 在 {start_date.date()} 之前的資料不足 {days_180} 天")
        
        # 取得序列（移除 NaN 值）
        rsi_series = df_before_start["RSI"].dropna()
        adx_series = df_before_start["ADX"].dropna()
        close_series = df_before_open["Close"]  # 使用開盤日期之前的收盤價
        
        # 取得最近的 5 天、30 天和 120 天序列（從最遠到最近）
        # 四捨五入到 1 位小數
        rsi_5 = [round(x, 1) for x in rsi_series.tail(days_5).tolist()]
        rsi_30 = [round(x, 1) for x in rsi_series.tail(days_30).tolist()]
        rsi_180 = [round(x, 1) for x in rsi_series.tail(days_180).tolist()]
        adx_5 = [round(x, 1) for x in adx_series.tail(days_5).tolist()]
        adx_30 = [round(x, 1) for x in adx_series.tail(days_30).tolist()]
        adx_180 = [round(x, 1) for x in adx_series.tail(days_180).tolist()]
        
        # 計算價格距離（使用昨日收盤價 = 開盤日期前一個交易日的收盤價）
        yesterday_close = close_series.iloc[-1]  # 開盤日期前一個交易日的收盤價
        yesterday_date = close_series.index[-1]  # 昨日的日期
        
        print(f"  昨日日期: {yesterday_date.strftime('%Y-%m-%d')} (美國時間)")
        
        price_dist_5 = calculate_price_distance(close_series, yesterday_close, 5)
        price_dist_30 = calculate_price_distance(close_series, yesterday_close, 30)
        price_dist_180 = calculate_price_distance(close_series, yesterday_close, 120)
        
        return {
            "RSI_5天": rsi_5,
            "RSI_30天": rsi_30,
            "RSI_180天": rsi_180,
            "ADX_5天": adx_5,
            "ADX_30天": adx_30,
            "ADX_180天": adx_180,
            "價格距離_5日": price_dist_5,
            "價格距離_30日": price_dist_30,
            "價格距離_180日": price_dist_180,
            "昨日收盤價": round(yesterday_close, 2),
            "實際資料天數": len(rsi_series)
        }
    
    except Exception as e:
        print(f"  處理 {ticker} 時發生錯誤: {str(e)}")
        return None

if __name__ == "__main__":
    # 執行主程式
    input_file = "量化交易.xlsx"
    sheet_name = "資料庫"
    
    print("="*60)
    print("美股技術指標計算系統 - RSI & ADX")
    print("="*60)
    print("\n計算指標：")
    print("  ✓ RSI (相對強弱指標) - 5天、30天和6個月序列")
    print("  ✓ ADX (平均趨向指標) - 5天、30天和6個月序列")
    print("  ✓ 價格距離 - 昨日收盤價距離過去高低點的百分比")
    print("  ✓ 盤中價格 - 開盤價、10分鐘最低價、1.5小時最高價等")
    print("    ⚠ 盤中數據僅限最近 7 天（yfinance 免費版限制）")
    print("\n資料來源：Yahoo Finance (免費)")
    print("="*60 + "\n")
    
    # 讀取 Excel
    print(f"正在讀取 {input_file} 的 '{sheet_name}' 工作表...")
    df = pd.read_excel(input_file, sheet_name=sheet_name)
    
    print(f"Excel 欄位: {df.columns.tolist()}")
    print(f"共有 {len(df)} 筆資料\n")
    
    # 使用正確的欄位名稱（支援多種格式）
    ticker_col = '公司代碼'
    
    # 嘗試找到日期欄位（支援多種名稱）
    date_col = None
    possible_date_cols = ['開盤日期(台灣時間)', '開盤日期', '日期', '交易日期']
    for col in possible_date_cols:
        if col in df.columns:
            date_col = col
            break
    
    # 檢查欄位是否存在
    if ticker_col not in df.columns:
        print(f"錯誤: 找不到 '{ticker_col}' 欄位")
        print(f"現有欄位: {df.columns.tolist()}")
        exit(1)
    
    if date_col is None:
        print(f"錯誤: 找不到日期欄位")
        print(f"支援的日期欄位名稱: {possible_date_cols}")
        print(f"現有欄位: {df.columns.tolist()}")
        exit(1)
    
    print(f"使用欄位: 股票代碼='{ticker_col}', 日期='{date_col}'\n")
    
    # 更新對應的欄位
    rsi_5_col = '5天 RSI 序列'
    rsi_30_col = '1個月 RSI 序列'
    rsi_180_col = '6個月 RSI 序列'
    adx_5_col = '5天 ADX 序列'
    adx_30_col = '1個月 ADX 序列'
    adx_180_col = '6個月 ADX 序列'
    
    # 價格距離欄位（可選）
    price_dist_cols = {
        '5日高價距離 (%)': ('價格距離_5日', '距離最高價(%)'),
        '5日低價距離 (%)': ('價格距離_5日', '距離最低價(%)'),
        '1個月高價距離 (%)': ('價格距離_30日', '距離最高價(%)'),
        '1個月低價距離 (%)': ('價格距離_30日', '距離最低價(%)'),
        '6個月高價距離 (%)': ('價格距離_180日', '距離最高價(%)'),
        '6個月低價距離 (%)': ('價格距離_180日', '距離最低價(%)'),
        '*昨日收盤價': ('昨日收盤價', None),
        '昨日收盤價': ('昨日收盤價', None)  # 支援兩種名稱
    }
    
    # 盤中價格欄位（可選）
    intraday_price_cols = {
        '*開盤價格': ('盤中價格', '開盤價'),
        '開盤價格': ('盤中價格', '開盤價'),
        '*10分鐘最低價': ('盤中價格', '10分鐘最低價'),
        '10分鐘最低價': ('盤中價格', '10分鐘最低價'),
        '*1.5小時最高價': ('盤中價格', '1.5小時最高價'),
        '1.5小時最高價': ('盤中價格', '1.5小時最高價'),
        '*最高價前的最低價': ('盤中價格', '最高價前的最低價'),
        '最高價前的最低價': ('盤中價格', '最高價前的最低價')
    }
    
    # 找到欄位索引
    col_indices = {}
    for col_name in [rsi_5_col, rsi_30_col, rsi_180_col, adx_5_col, adx_30_col, adx_180_col]:
        if col_name in df.columns:
            col_indices[col_name] = df.columns.get_loc(col_name) + 1
    
    # 找到價格距離欄位索引
    for col_name in price_dist_cols.keys():
        if col_name in df.columns:
            col_indices[col_name] = df.columns.get_loc(col_name) + 1
    
    # 找到盤中價格欄位索引
    for col_name in intraday_price_cols.keys():
        if col_name in df.columns:
            col_indices[col_name] = df.columns.get_loc(col_name) + 1
    
    # 統計變數
    processed_count = 0
    skipped_count = 0
    failed_count = 0
    
    # 儲存需要更新的資料
    updates = {}
    
    # 逐行處理
    for idx, row in df.iterrows():
        ticker = row[ticker_col]
        date = row[date_col]
        
        # 跳過空值
        if pd.isna(ticker) or pd.isna(date):
            print(f"跳過第 {idx + 1} 筆: 資料不完整")
            skipped_count += 1
            continue
        
        # 檢查 RSI/ADX 是否已經計算過
        has_rsi_5 = not pd.isna(row[rsi_5_col]) and str(row[rsi_5_col]).strip() not in ['', '[]', 'nan']
        has_rsi_30 = not pd.isna(row[rsi_30_col]) and str(row[rsi_30_col]).strip() not in ['', '[]', 'nan']
        has_rsi_180 = not pd.isna(row[rsi_180_col]) and str(row[rsi_180_col]).strip() not in ['', '[]', 'nan']
        has_adx_5 = not pd.isna(row[adx_5_col]) and str(row[adx_5_col]).strip() not in ['', '[]', 'nan']
        has_adx_30 = not pd.isna(row[adx_30_col]) and str(row[adx_30_col]).strip() not in ['', '[]', 'nan']
        has_adx_180 = not pd.isna(row[adx_180_col]) and str(row[adx_180_col]).strip() not in ['', '[]', 'nan']
        
        # 檢查是否需要計算 RSI/ADX
        need_rsi_adx = not (has_rsi_5 and has_rsi_30 and has_rsi_180 and has_adx_5 and has_adx_30 and has_adx_180)
        
        # 檢查是否需要計算價格距離（檢查昨日收盤價欄位）
        need_price_dist = False
        yesterday_close_col = '*昨日收盤價' if '*昨日收盤價' in df.columns else '昨日收盤價'
        if yesterday_close_col in df.columns:
            has_yesterday_close = not pd.isna(row[yesterday_close_col]) and str(row[yesterday_close_col]).strip() not in ['', 'nan']
            need_price_dist = not has_yesterday_close
        
        # 檢查是否需要計算盤中數據（檢查開盤價欄位）
        need_intraday = False
        open_price_col = '*開盤價格' if '*開盤價格' in df.columns else '開盤價格'
        if open_price_col in df.columns:
            has_open_price = not pd.isna(row[open_price_col]) and str(row[open_price_col]).strip() not in ['', 'nan']
            need_intraday = not has_open_price
        
        # 如果所有資料都已經有了，跳過
        if not need_rsi_adx and not need_price_dist and not need_intraday:
            print(f"跳過第 {idx + 1} 筆: {ticker} (日期: {date}) - 所有資料已完整")
            skipped_count += 1
            continue
        
        print(f"處理第 {idx + 1} 筆: {ticker} (日期: {date})")
        
        # 初始化結果
        result = None
        intraday_data = None
        
        # 只在需要時計算 RSI/ADX 和價格距離
        if need_rsi_adx or need_price_dist:
            result = calculate_rsi_adx_sequences(ticker, date)
            if not result:
                print(f"  ✗ 無法獲取股價資料")
                failed_count += 1
                print()
                continue
        
        # 只在需要時計算盤中數據
        if need_intraday:
            intraday_data = calculate_intraday_prices(ticker, date)
        
        # 儲存更新資料
        excel_row = idx + 2
        updates[excel_row] = {}
        
        # 只在需要時更新 RSI/ADX
        if need_rsi_adx and result and len(result["RSI_5天"]) > 0:
            updates[excel_row].update({
                rsi_5_col: str(result["RSI_5天"]),
                rsi_30_col: str(result["RSI_30天"]),
                rsi_180_col: str(result["RSI_180天"]),
                adx_5_col: str(result["ADX_5天"]),
                adx_30_col: str(result["ADX_30天"]),
                adx_180_col: str(result["ADX_180天"])
            })
        
        # 只在需要時更新價格距離
        if need_price_dist and result:
            for col_name, (result_key, sub_key) in price_dist_cols.items():
                if col_name in col_indices and result_key in result:
                    if sub_key:
                        if result[result_key]:
                            updates[excel_row][col_name] = result[result_key][sub_key]
                    else:
                        updates[excel_row][col_name] = result[result_key]
        
        # 只在需要時更新盤中價格
        if need_intraday and intraday_data:
            for col_name, (result_key, sub_key) in intraday_price_cols.items():
                if col_name in col_indices and sub_key in intraday_data:
                    if intraday_data[sub_key] is not None:
                        updates[excel_row][col_name] = intraday_data[sub_key]
        
        # 顯示處理結果
        if updates[excel_row]:  # 如果有任何更新
            print(f"  ✓ 完成")
            
            if need_rsi_adx and result:
                print(f"    - RSI/ADX 已更新 (實際資料: {result['實際資料天數']} 天)")
                if len(result['RSI_5天']) >= 3:
                    print(f"    - RSI 5天前3筆: {result['RSI_5天'][:3]}")
            
            if need_price_dist and result:
                if '價格距離_5日' in result and result['價格距離_5日']:
                    print(f"    - 昨日收盤價: ${result['昨日收盤價']}")
                    print(f"    - 5日距離: 高 {result['價格距離_5日']['距離最高價(%)']}%, 低 {result['價格距離_5日']['距離最低價(%)']}%")
            
            if need_intraday and intraday_data:
                print(f"    - 盤中數據: 開盤 ${intraday_data['開盤價']}, 10分鐘低 ${intraday_data['10分鐘最低價']}")
            
            processed_count += 1
        else:
            print(f"  ⚠ 無可用資料")
            failed_count += 1
        
        print()
    
    # 顯示統計資訊
    print(f"\n{'='*60}")
    print(f"處理完成統計:")
    print(f"  新計算: {processed_count} 筆")
    print(f"  已跳過: {skipped_count} 筆（已有資料或資料不完整）")
    print(f"  失敗: {failed_count} 筆")
    print(f"  總計: {len(df)} 筆")
    print(f"{'='*60}\n")
    
    # 如果有需要更新的資料，使用 openpyxl 直接寫入
    if updates:
        print(f"正在更新 {input_file} 的 '{sheet_name}' 工作表（保留原有格式）...")
        
        from openpyxl import load_workbook
        from copy import copy
        
        wb = load_workbook(input_file)
        ws = wb[sheet_name]
        
        # 更新儲存格
        for excel_row, data in updates.items():
            for col_name, value in data.items():
                if col_name in col_indices:
                    col_idx = col_indices[col_name]
                    cell = ws.cell(row=excel_row, column=col_idx)
                    
                    # 策略：尋找同一欄位中已有手動輸入資料的儲存格作為格式參考
                    # 如果找不到，則從同一行的「公司代碼」欄位複製格式
                    reference_cell = None
                    
                    # 先嘗試在同一欄位中找到已有資料的儲存格（往上找最多 20 行）
                    for ref_row in range(max(2, excel_row - 20), excel_row):
                        ref_cell = ws.cell(row=ref_row, column=col_idx)
                        if ref_cell.value is not None and str(ref_cell.value).strip() not in ['', 'nan']:
                            reference_cell = ref_cell
                            break
                    
                    # 如果找不到，使用同一行的「公司代碼」欄位
                    if reference_cell is None:
                        ticker_col_idx = df.columns.get_loc('公司代碼') + 1
                        reference_cell = ws.cell(row=excel_row, column=ticker_col_idx)
                    
                    # 使用 copy() 完整複製參考格式（包括顏色、粗體等所有屬性）
                    if reference_cell.font:
                        cell.font = copy(reference_cell.font)
                    if reference_cell.alignment:
                        cell.alignment = copy(reference_cell.alignment)
                    if reference_cell.border:
                        cell.border = copy(reference_cell.border)
                    if reference_cell.fill:
                        cell.fill = copy(reference_cell.fill)
                    if reference_cell.number_format:
                        cell.number_format = reference_cell.number_format
                    
                    # 最後更新值（在設定格式之後）
                    cell.value = value
        
        # 儲存工作簿
        wb.save(input_file)
        wb.close()
        
        print("✓ 完成！格式已保留。")
    else:
        print("沒有需要更新的資料。")
    
    print("\n提示：")
    print("  - 序列排序為從最遠到最近 [第N天前, ..., 第1天前]")
    print("  - RSI > 70: 超買，RSI < 30: 超賣")
    print("  - ADX > 25: 強趨勢，ADX < 20: 弱趨勢")
    print("  - 6個月序列約包含 120 個交易日的資料")
