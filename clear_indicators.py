"""
清除 Excel 中的 RSI/ADX 資料，以便重新計算
"""

import pandas as pd
from openpyxl import load_workbook

input_file = "量化交易.xlsx"
sheet_name = "資料庫"

print("正在清除 RSI/ADX 資料...")

# 讀取 Excel
df = pd.read_excel(input_file, sheet_name=sheet_name)

# 要清除的欄位
columns_to_clear = [
    '5天 RSI 序列',
    '5天 ADX 序列',
    '1個月 RSI 序列',
    '1個月 ADX 序列',
    '6個月 RSI 序列',
    '6個月 ADX 序列'
]

# 找到欄位索引
col_indices = {}
for col_name in columns_to_clear:
    if col_name in df.columns:
        col_indices[col_name] = df.columns.get_loc(col_name) + 1
        print(f"  找到欄位: {col_name}")

if not col_indices:
    print("✗ 找不到任何 RSI/ADX 欄位")
    exit(1)

# 使用 openpyxl 清除資料
wb = load_workbook(input_file)
ws = wb[sheet_name]

cleared_count = 0
for row_idx in range(2, len(df) + 2):  # 從第 2 行開始（第 1 行是標題）
    for col_name, col_idx in col_indices.items():
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value:
            cell.value = None
            cleared_count += 1

# 儲存
wb.save(input_file)
wb.close()

print(f"\n✓ 已清除 {cleared_count} 個儲存格")
print(f"✓ 共處理 {len(df)} 筆資料")
print("\n現在可以執行 python calculate_indicators.py 重新計算")
