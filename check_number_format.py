"""
檢查數字格式
"""

from openpyxl import load_workbook
import pandas as pd

input_file = "量化交易.xlsx"
sheet_name = "資料庫"

print("="*60)
print("檢查數字格式")
print("="*60)

# 讀取 Excel
df = pd.read_excel(input_file, sheet_name=sheet_name)

# 載入工作簿
wb = load_workbook(input_file)
ws = wb[sheet_name]

# 檢查第 2 行的數字格式
print("\n檢查第 2 行的數字格式:")
print("-"*60)

# 檢查幾個數字欄位
test_cols = {
    '5日高價距離 (%)': '手動輸入',
    '*昨日收盤價': '自動填入',
    '*開盤價格': '自動填入',
    '產業': '手動輸入（數字）'
}

for col_name, col_type in test_cols.items():
    if col_name in df.columns:
        col_idx = df.columns.get_loc(col_name) + 1
        cell = ws.cell(row=2, column=col_idx)
        
        print(f"\n{col_name} ({col_type}):")
        print(f"  值: {cell.value}")
        print(f"  值類型: {type(cell.value).__name__}")
        print(f"  數字格式: {cell.number_format}")
        print(f"  字型: {cell.font.name} {cell.font.size}pt")

wb.close()

print("\n" + "="*60)
print("說明:")
print("  - 'General' = 一般格式")
print("  - '0.00' = 兩位小數")
print("  - '0.0' = 一位小數")
print("  - 如果數字格式不一致，可能導致顯示不同")
print("="*60)
