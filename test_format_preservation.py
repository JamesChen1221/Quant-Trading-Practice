"""
測試格式保留功能
"""

from openpyxl import load_workbook
from copy import copy

input_file = "量化交易.xlsx"
sheet_name = "資料庫"

print("="*60)
print("測試格式保留功能")
print("="*60)

# 載入工作簿
wb = load_workbook(input_file)
ws = wb[sheet_name]

# 檢查第 2 行（第一筆資料）的格式
print("\n檢查第 2 行的格式:")
row = 2

# 檢查幾個欄位的格式
test_cols = ['A', 'B', 'C']  # 前幾個欄位
for col in test_cols:
    cell = ws[f'{col}{row}']
    print(f"\n欄位 {col}{row}:")
    print(f"  值: {cell.value}")
    print(f"  字型: {cell.font.name}, 大小: {cell.font.size}, 粗體: {cell.font.bold}")
    print(f"  對齊: 水平={cell.alignment.horizontal}, 垂直={cell.alignment.vertical}")
    print(f"  數字格式: {cell.number_format}")

# 測試寫入並保留格式
print("\n" + "="*60)
print("測試寫入資料並保留格式")
print("="*60)

test_cell = ws['A2']
print(f"\n原始格式:")
print(f"  字型: {test_cell.font.name}, 大小: {test_cell.font.size}")
print(f"  原始值: {test_cell.value}")

# 保存格式
original_font = copy(test_cell.font)
original_alignment = copy(test_cell.alignment)
original_border = copy(test_cell.border)
original_fill = copy(test_cell.fill)
original_number_format = test_cell.number_format

# 修改值
old_value = test_cell.value
test_cell.value = "測試值"

# 恢復格式
test_cell.font = original_font
test_cell.alignment = original_alignment
test_cell.border = original_border
test_cell.fill = original_fill
test_cell.number_format = original_number_format

print(f"\n修改後:")
print(f"  字型: {test_cell.font.name}, 大小: {test_cell.font.size}")
print(f"  新值: {test_cell.value}")

# 恢復原值（不儲存）
test_cell.value = old_value

print("\n✓ 格式保留測試完成（未儲存變更）")
print("\n說明:")
print("  程式會在寫入新值時保留以下格式:")
print("  - 字型（名稱、大小、粗體、顏色等）")
print("  - 對齊方式")
print("  - 邊框")
print("  - 背景色")
print("  - 數字格式")

wb.close()
