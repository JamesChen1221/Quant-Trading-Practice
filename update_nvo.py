"""
更新 NVO 的「最高價前的最低價」
"""

import pandas as pd
from openpyxl import load_workbook
from copy import copy

input_file = "量化交易.xlsx"
sheet_name = "資料庫"

print("正在更新 NVO 的「最高價前的最低價」...")

# 讀取 Excel
df = pd.read_excel(input_file, sheet_name=sheet_name)

# 找到 NVO 的行
nvo_rows = df[df['公司代碼'] == 'NVO'].index.tolist()

if not nvo_rows:
    print("✗ 找不到 NVO")
    exit(1)

print(f"找到 {len(nvo_rows)} 筆 NVO 資料")

# 使用 openpyxl 更新
wb = load_workbook(input_file)
ws = wb[sheet_name]

# 找到「最高價前的最低價」欄位
low_before_high_col = '*最高價前的最低價'
if low_before_high_col not in df.columns:
    low_before_high_col = '最高價前的最低價'

if low_before_high_col not in df.columns:
    print("✗ 找不到「最高價前的最低價」欄位")
    wb.close()
    exit(1)

col_idx = df.columns.get_loc(low_before_high_col) + 1

# 更新每一筆 NVO
for idx in nvo_rows:
    excel_row = idx + 2  # Excel 行號從 2 開始
    cell = ws.cell(row=excel_row, column=col_idx)
    
    # 保存格式
    original_font = copy(cell.font)
    original_alignment = copy(cell.alignment)
    original_border = copy(cell.border)
    original_fill = copy(cell.fill)
    original_number_format = cell.number_format
    
    # 設為 58.61（最高價）
    cell.value = 58.61
    
    # 恢復格式
    cell.font = original_font
    cell.alignment = original_alignment
    cell.border = original_border
    cell.fill = original_fill
    cell.number_format = original_number_format
    
    print(f"  ✓ 更新第 {excel_row} 行: ${cell.value}")

# 儲存
wb.save(input_file)
wb.close()

print("\n✓ 更新完成！")
print("\n說明:")
print("  NVO 的最高價出現在開盤後第 11 分鐘")
print("  「最高價前的最低價」設為最高價 $58.61")
print("  表示開盤後立即達到高點，之後不再向上")
